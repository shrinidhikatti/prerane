# core/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.views import LoginView, LogoutView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import Group
from django.views.generic import CreateView, UpdateView, DeleteView, ListView, DetailView, TemplateView
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db import transaction
import openpyxl
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from .mixin import *
from .models import *
from .forms import *

class CustomLoginView(LoginView):
    template_name = 'registration/login.html'
    
    def get_success_url(self):
        return reverse_lazy('dashboard')

class DashboardView(LoginRequiredMixin, RoleContextMixin, TemplateView):
    template_name = 'core/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        try:
            if user.groups.filter(name='DDPI').exists():
                ddpi_profile = DDPIProfile.objects.select_related('district').get(user=user)
                context.update({
                    'talukas_count': Taluka.objects.filter(district=ddpi_profile.district).count(),
                    'beos_count': BEOProfile.objects.filter(taluka__district=ddpi_profile.district).count(),
                    'subjects_count': Subject.objects.count(),
                    'assignments_count': Assignment.objects.count(),
                })
                # Get task statistics for district
                context.update(self.get_task_statistics_for_district(ddpi_profile.district))
                
            elif user.groups.filter(name='BEO').exists():
                beo_profile = BEOProfile.objects.select_related('taluka').get(user=user)
                context.update({
                    'schools_count': School.objects.filter(taluka=beo_profile.taluka).count(),
                    'principals_count': PrincipalProfile.objects.filter(school__taluka=beo_profile.taluka).count(),
                })
                # Get task statistics for taluka
                context.update(self.get_task_statistics_for_taluka(beo_profile.taluka))
                
            elif user.groups.filter(name='Principal').exists():
                principal_profile = PrincipalProfile.objects.select_related('school').get(user=user)
                context.update({
                    'students_count': Student.objects.filter(school=principal_profile.school).count(),
                    'assignments_count': Assignment.objects.count(),
                })
                # Get task statistics for school
                context.update(self.get_task_statistics_for_school(principal_profile.school))
                
        except:
            # Handle case where profile doesn't exist
            pass
            
        return context
    
    def get_task_statistics_for_district(self, district):
        students = Student.objects.filter(school__taluka__district=district).select_related('school', 'school__taluka')
        return self.calculate_task_statistics(students)
    
    def get_task_statistics_for_taluka(self, taluka):
        students = Student.objects.filter(school__taluka=taluka).select_related('school')
        return self.calculate_task_statistics(students)
    
    def get_task_statistics_for_school(self, school):
        students = Student.objects.filter(school=school).select_related('school')
        return self.calculate_task_statistics(students)
    
    def calculate_task_statistics(self, students):
        from django.db.models import Count, Q
        from collections import defaultdict
        
        # Get all assignments that have tasks for the standards in our student set
        student_standards = set(students.values_list('standard', flat=True))
        if not student_standards:
            return {
                'total_tasks': 0,
                'solved_tasks': 0,
                'unsolved_tasks': 0,
                'unassigned_tasks': 0,
                'solved_percentage': 0,
                'unsolved_percentage': 0,
                'unassigned_percentage': 0,
            }
        
        # Get assignments with optimized query
        assignments = Assignment.objects.filter(
            standard__in=student_standards
        ).select_related('subject')
        
        if not assignments:
            return {
                'total_tasks': 0,
                'solved_tasks': 0,
                'unsolved_tasks': 0,
                'unassigned_tasks': 0,
                'solved_percentage': 0,
                'unsolved_percentage': 0,
                'unassigned_percentage': 0,
            }
        
        # Calculate total possible tasks efficiently using group counting
        students_by_standard = defaultdict(int)
        for standard in students.values('standard').annotate(count=Count('id')):
            students_by_standard[standard['standard']] = standard['count']
        
        total_tasks = 0
        for assignment in assignments:
            student_count = students_by_standard.get(assignment.standard, 0)
            total_tasks += student_count * len(assignment.tasks)
        
        if total_tasks == 0:
            return {
                'total_tasks': 0,
                'solved_tasks': 0,
                'unsolved_tasks': 0,
                'unassigned_tasks': 0,
                'solved_percentage': 0,
                'unsolved_percentage': 0,
                'unassigned_percentage': 0,
            }
        
        # Get all evaluations for these students and assignments in one optimized query
        student_ids = list(students.values_list('id', flat=True))
        assignment_ids = list(assignments.values_list('id', flat=True))
        
        # Use aggregate query for better performance
        evaluation_counts = TaskEvaluation.objects.filter(
            student_id__in=student_ids,
            assignment_id__in=assignment_ids
        ).aggregate(
            solved_count=Count('id', filter=Q(status='solved')),
            unsolved_count=Count('id', filter=Q(status='unsolved'))
        )
        
        solved_tasks = evaluation_counts['solved_count'] or 0
        unsolved_tasks = evaluation_counts['unsolved_count'] or 0
        total_evaluated = solved_tasks + unsolved_tasks
        unassigned_tasks = total_tasks - total_evaluated
        
        return {
            'total_tasks': total_tasks,
            'solved_tasks': solved_tasks,
            'unsolved_tasks': unsolved_tasks,
            'unassigned_tasks': unassigned_tasks,
            'solved_percentage': round((solved_tasks / total_tasks * 100) if total_tasks > 0 else 0, 1),
            'unsolved_percentage': round((unsolved_tasks / total_tasks * 100) if total_tasks > 0 else 0, 1),
            'unassigned_percentage': round((unassigned_tasks / total_tasks * 100) if total_tasks > 0 else 0, 1),
        }
      
# DDPI Views
class ManageTalukaListView(LoginRequiredMixin, DDPIRequiredMixin, RoleContextMixin, ListView):
    model = Taluka
    template_name = 'core/ddpi/taluka_list.html'
    context_object_name = 'talukas'
    
    def get_queryset(self):
        ddpi_profile = DDPIProfile.objects.get(user=self.request.user)
        return Taluka.objects.filter(district=ddpi_profile.district)

class ManageTalukaCreateView(LoginRequiredMixin, DDPIRequiredMixin, RoleContextMixin, CreateView):
    model = Taluka
    form_class = TalukaForm
    template_name = 'core/ddpi/taluka_form.html'
    success_url = reverse_lazy('manage_talukas')
    
    def get_form(self):
        form = super().get_form()
        ddpi_profile = DDPIProfile.objects.get(user=self.request.user)
        form.fields['district'].queryset = District.objects.filter(id=ddpi_profile.district.id)
        form.fields['district'].initial = ddpi_profile.district
        return form
    
    def form_valid(self, form):
        messages.success(self.request, 'Taluka created successfully.')
        return super().form_valid(form)

class ManageTalukaUpdateView(LoginRequiredMixin, DDPIRequiredMixin, RoleContextMixin, UpdateView):
    model = Taluka
    form_class = TalukaForm
    template_name = 'core/ddpi/taluka_form.html'
    success_url = reverse_lazy('manage_talukas')
    
    def get_queryset(self):
        ddpi_profile = DDPIProfile.objects.get(user=self.request.user)
        return Taluka.objects.filter(district=ddpi_profile.district)
    
    def form_valid(self, form):
        messages.success(self.request, 'Taluka updated successfully.')
        return super().form_valid(form)

class ManageTalukaDeleteView(LoginRequiredMixin, DDPIRequiredMixin, RoleContextMixin, DeleteView):
    model = Taluka
    template_name = 'core/ddpi/taluka_confirm_delete.html'
    success_url = reverse_lazy('manage_talukas')
    
    def get_queryset(self):
        ddpi_profile = DDPIProfile.objects.get(user=self.request.user)
        return Taluka.objects.filter(district=ddpi_profile.district)
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Taluka deleted successfully.')
        return super().delete(request, *args, **kwargs)

class ManageSubjectListView(LoginRequiredMixin, DDPIRequiredMixin, RoleContextMixin, ListView):
    model = Subject
    template_name = 'core/ddpi/subject_list.html'
    context_object_name = 'subjects'

class ManageSubjectCreateView(LoginRequiredMixin, DDPIRequiredMixin, RoleContextMixin, CreateView):
    model = Subject
    form_class = SubjectForm
    template_name = 'core/ddpi/subject_form.html'
    success_url = reverse_lazy('manage_subjects')
    
    def form_valid(self, form):
        messages.success(self.request, 'Subject created successfully.')
        return super().form_valid(form)

class ManageSubjectUpdateView(LoginRequiredMixin, DDPIRequiredMixin, RoleContextMixin, UpdateView):
    model = Subject
    form_class = SubjectForm
    template_name = 'core/ddpi/subject_form.html'
    success_url = reverse_lazy('manage_subjects')
    
    def form_valid(self, form):
        messages.success(self.request, 'Subject updated successfully.')
        return super().form_valid(form)

class ManageSubjectDeleteView(LoginRequiredMixin, DDPIRequiredMixin, RoleContextMixin, DeleteView):
    model = Subject
    template_name = 'core/ddpi/subject_confirm_delete.html'
    success_url = reverse_lazy('manage_subjects')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Subject deleted successfully.')
        return super().delete(request, *args, **kwargs)

class ManageBEOListView(LoginRequiredMixin, DDPIRequiredMixin, RoleContextMixin, ListView):
    model = BEOProfile
    template_name = 'core/ddpi/beo_list.html'
    context_object_name = 'beos'
    
    def get_queryset(self):
        ddpi_profile = DDPIProfile.objects.get(user=self.request.user)
        return BEOProfile.objects.filter(taluka__district=ddpi_profile.district)

class ManageBEOCreateView(LoginRequiredMixin, DDPIRequiredMixin, RoleContextMixin, CreateView):
    form_class = BEOCreationForm
    template_name = 'core/ddpi/beo_form.html'
    success_url = reverse_lazy('manage_beos')
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        ddpi_profile = DDPIProfile.objects.get(user=self.request.user)
        kwargs['district'] = ddpi_profile.district
        return kwargs
    
    def form_valid(self, form):
        with transaction.atomic():
            user = form.save()
            beo_group = Group.objects.get(name='BEO')
            user.groups.add(beo_group)
            
            BEOProfile.objects.create(
                user=user,
                taluka=form.cleaned_data['taluka']
            )
            
        messages.success(self.request, 'BEO created successfully.')
        return super().form_valid(form)

class ManageBEOUpdateView(LoginRequiredMixin, DDPIRequiredMixin, RoleContextMixin, UpdateView):
    model = BEOProfile
    fields = ['taluka']
    template_name = 'core/ddpi/beo_form.html'
    success_url = reverse_lazy('manage_beos')
    
    def get_queryset(self):
        ddpi_profile = DDPIProfile.objects.get(user=self.request.user)
        return BEOProfile.objects.filter(taluka__district=ddpi_profile.district)
    
    def form_valid(self, form):
        messages.success(self.request, 'BEO updated successfully.')
        return super().form_valid(form)

class ManageBEODeleteView(LoginRequiredMixin, DDPIRequiredMixin, RoleContextMixin, DeleteView):
    model = BEOProfile
    template_name = 'core/ddpi/beo_confirm_delete.html'
    success_url = reverse_lazy('manage_beos')
    
    def get_queryset(self):
        ddpi_profile = DDPIProfile.objects.get(user=self.request.user)
        return BEOProfile.objects.filter(taluka__district=ddpi_profile.district)
    
    def delete(self, request, *args, **kwargs):
        beo = self.get_object()
        beo.user.delete()  # This will also delete the BEOProfile due to CASCADE
        messages.success(request, 'BEO deleted successfully.')
        return redirect(self.success_url)

class ManageAssignmentListView(LoginRequiredMixin, DDPIRequiredMixin, RoleContextMixin, ListView):
    model = Assignment
    template_name = 'core/ddpi/assignment_list.html'
    context_object_name = 'assignments'

class ManageAssignmentCreateView(LoginRequiredMixin, DDPIRequiredMixin, RoleContextMixin, CreateView):
    model = Assignment
    form_class = AssignmentForm
    template_name = 'core/ddpi/assignment_form.html'
    success_url = reverse_lazy('manage_assignments')
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'Assignment created successfully.')
        return super().form_valid(form)

class ManageAssignmentUpdateView(LoginRequiredMixin, DDPIRequiredMixin, RoleContextMixin, UpdateView):
    model = Assignment
    form_class = AssignmentForm
    template_name = 'core/ddpi/assignment_form.html'
    success_url = reverse_lazy('manage_assignments')
    
    def get_initial(self):
        initial = super().get_initial()
        if self.object.tasks:
            initial['tasks'] = '\n'.join(self.object.tasks)
        return initial
    
    def form_valid(self, form):
        messages.success(self.request, 'Assignment updated successfully.')
        return super().form_valid(form)

class ManageAssignmentDeleteView(LoginRequiredMixin, DDPIRequiredMixin, RoleContextMixin, DeleteView):
    model = Assignment
    template_name = 'core/ddpi/assignment_confirm_delete.html'
    success_url = reverse_lazy('manage_assignments')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Assignment deleted successfully.')
        return super().delete(request, *args, **kwargs)

# BEO Views
class ManageSchoolListView(LoginRequiredMixin, BEORequiredMixin, RoleContextMixin, ListView):
    model = School
    template_name = 'core/beo/school_list.html'
    context_object_name = 'schools'
    
    def get_queryset(self):
        beo_profile = BEOProfile.objects.get(user=self.request.user)
        return School.objects.filter(taluka=beo_profile.taluka)

class ManageSchoolCreateView(LoginRequiredMixin, BEORequiredMixin, RoleContextMixin, CreateView):
    model = School
    form_class = SchoolForm
    template_name = 'core/beo/school_form.html'
    success_url = reverse_lazy('manage_schools')
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        beo_profile = BEOProfile.objects.get(user=self.request.user)
        kwargs['taluka'] = beo_profile.taluka
        return kwargs
    
    def form_valid(self, form):
        messages.success(self.request, 'School created successfully.')
        return super().form_valid(form)

class ManageSchoolUpdateView(LoginRequiredMixin, BEORequiredMixin, RoleContextMixin, UpdateView):
    model = School
    form_class = SchoolForm
    template_name = 'core/beo/school_form.html'
    success_url = reverse_lazy('manage_schools')
    
    def get_queryset(self):
        beo_profile = BEOProfile.objects.get(user=self.request.user)
        return School.objects.filter(taluka=beo_profile.taluka)
    
    def form_valid(self, form):
        messages.success(self.request, 'School updated successfully.')
        return super().form_valid(form)

class ManageSchoolDeleteView(LoginRequiredMixin, BEORequiredMixin, RoleContextMixin, DeleteView):
    model = School
    template_name = 'core/beo/school_confirm_delete.html'
    success_url = reverse_lazy('manage_schools')
    
    def get_queryset(self):
        beo_profile = BEOProfile.objects.get(user=self.request.user)
        return School.objects.filter(taluka=beo_profile.taluka)
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'School deleted successfully.')
        return super().delete(request, *args, **kwargs)

class ManagePrincipalListView(LoginRequiredMixin, BEORequiredMixin, RoleContextMixin, ListView):
    model = PrincipalProfile
    template_name = 'core/beo/principal_list.html'
    context_object_name = 'principals'
    
    def get_queryset(self):
        beo_profile = BEOProfile.objects.get(user=self.request.user)
        return PrincipalProfile.objects.filter(school__taluka=beo_profile.taluka)

class ManagePrincipalCreateView(LoginRequiredMixin, BEORequiredMixin, RoleContextMixin, CreateView):
    form_class = PrincipalCreationForm
    template_name = 'core/beo/principal_form.html'
    success_url = reverse_lazy('manage_principals')
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        beo_profile = BEOProfile.objects.get(user=self.request.user)
        kwargs['taluka'] = beo_profile.taluka
        return kwargs
    
    def form_valid(self, form):
        with transaction.atomic():
            user = form.save()
            principal_group = Group.objects.get(name='Principal')
            user.groups.add(principal_group)
            
            PrincipalProfile.objects.create(
                user=user,
                school=form.cleaned_data['school']
            )
            
        messages.success(self.request, 'Principal created successfully.')
        return super().form_valid(form)

class ManagePrincipalUpdateView(LoginRequiredMixin, BEORequiredMixin, RoleContextMixin, UpdateView):
    model = PrincipalProfile
    form_class = PrincipalUpdateForm
    template_name = 'core/beo/principal_form.html'
    success_url = reverse_lazy('manage_principals')
    
    def get_queryset(self):
        beo_profile = BEOProfile.objects.get(user=self.request.user)
        return PrincipalProfile.objects.filter(school__taluka=beo_profile.taluka)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        beo_profile = BEOProfile.objects.get(user=self.request.user)
        kwargs['taluka'] = beo_profile.taluka
        return kwargs
    
    def form_valid(self, form):
        response = super().form_valid(form)
        new_password = form.cleaned_data.get('new_password')
        if new_password:
            principal_user = form.instance.user
            principal_user.set_password(new_password)
            principal_user.save()
            messages.success(self.request, 'Principal updated successfully with new password.')
        else:
            messages.success(self.request, 'Principal updated successfully.')
        return response

class ManagePrincipalDeleteView(LoginRequiredMixin, BEORequiredMixin, RoleContextMixin, DeleteView):
    model = PrincipalProfile
    template_name = 'core/beo/principal_confirm_delete.html'
    success_url = reverse_lazy('manage_principals')
    
    def get_queryset(self):
        beo_profile = BEOProfile.objects.get(user=self.request.user)
        return PrincipalProfile.objects.filter(school__taluka=beo_profile.taluka)
    
    def delete(self, request, *args, **kwargs):
        principal = self.get_object()
        principal.user.delete()  # This will also delete the PrincipalProfile due to CASCADE
        messages.success(request, 'Principal deleted successfully.')
        return redirect(self.success_url)

# Principal Views
class ManageStudentListView(LoginRequiredMixin, PrincipalRequiredMixin, RoleContextMixin, ListView):
    model = Student
    template_name = 'core/principal/student_list.html'
    context_object_name = 'students'
    
    def get_queryset(self):
        principal_profile = PrincipalProfile.objects.get(user=self.request.user)
        return Student.objects.filter(school=principal_profile.school)

class ManageStudentCreateView(LoginRequiredMixin, PrincipalRequiredMixin, RoleContextMixin, CreateView):
    model = Student
    form_class = StudentForm
    template_name = 'core/principal/student_form.html'
    success_url = reverse_lazy('manage_students')
    
    def form_valid(self, form):
        principal_profile = PrincipalProfile.objects.get(user=self.request.user)
        form.instance.school = principal_profile.school
        messages.success(self.request, 'Student created successfully.')
        return super().form_valid(form)

class ManageStudentUpdateView(LoginRequiredMixin, PrincipalRequiredMixin, RoleContextMixin, UpdateView):
    model = Student
    form_class = StudentForm
    template_name = 'core/principal/student_form.html'
    success_url = reverse_lazy('manage_students')
    
    def get_queryset(self):
        principal_profile = PrincipalProfile.objects.get(user=self.request.user)
        return Student.objects.filter(school=principal_profile.school)
    
    def form_valid(self, form):
        messages.success(self.request, 'Student updated successfully.')
        return super().form_valid(form)

class ManageStudentDeleteView(LoginRequiredMixin, PrincipalRequiredMixin, RoleContextMixin, DeleteView):
    model = Student
    template_name = 'core/principal/student_confirm_delete.html'
    success_url = reverse_lazy('manage_students')
    
    def get_queryset(self):
        principal_profile = PrincipalProfile.objects.get(user=self.request.user)
        return Student.objects.filter(school=principal_profile.school)
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Student deleted successfully.')
        return super().delete(request, *args, **kwargs)

class EvaluateAssignmentListView(LoginRequiredMixin, PrincipalRequiredMixin, RoleContextMixin, ListView):
    model = Assignment
    template_name = 'core/principal/assignment_evaluation_list.html'
    context_object_name = 'assignments'

class EvaluateAssignmentView(LoginRequiredMixin, PrincipalRequiredMixin, RoleContextMixin, TemplateView):
    template_name = 'core/principal/assignment_evaluation.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        assignment = get_object_or_404(Assignment, pk=kwargs['pk'])
        principal_profile = PrincipalProfile.objects.get(user=self.request.user)
        students = Student.objects.filter(
            school=principal_profile.school,
            standard=assignment.standard
        )
        
        context['assignment'] = assignment
        context['students'] = students
        context['tasks'] = assignment.tasks
        
        # Get existing evaluations
        evaluations = {}
        for student in students:
            evaluations[student.id] = {}
            for i in range(len(assignment.tasks)):
                try:
                    eval_obj = TaskEvaluation.objects.get(
                        student=student,
                        assignment=assignment,
                        task_index=i
                    )
                    evaluations[student.id][i] = eval_obj.status
                except TaskEvaluation.DoesNotExist:
                    evaluations[student.id][i] = 'unsolved'
        
        context['evaluations'] = evaluations
        return context
    
    def post(self, request, *args, **kwargs):
        assignment = get_object_or_404(Assignment, pk=kwargs['pk'])
        principal_profile = PrincipalProfile.objects.get(user=request.user)
        students = Student.objects.filter(
            school=principal_profile.school,
            standard=assignment.standard
        )
        
        with transaction.atomic():
            for student in students:
                for i in range(len(assignment.tasks)):
                    field_name = f'student_{student.id}_task_{i}'
                    status = request.POST.get(field_name, 'unsolved')
                    
                    TaskEvaluation.objects.update_or_create(
                        student=student,
                        assignment=assignment,
                        task_index=i,
                        defaults={
                            'status': status,
                            'evaluated_by': request.user
                        }
                    )
        
        messages.success(request, 'Assignment evaluation updated successfully.')
        return redirect('evaluate_assignment', pk=assignment.pk)
    

class PasswordChangeView(LoginRequiredMixin, RoleContextMixin, TemplateView):
    template_name = 'core/password_change.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = PasswordChangeForm(user=self.request.user)
        return context
    
    def post(self, request, *args, **kwargs):
        form = PasswordChangeForm(request.POST, user=request.user)
        if form.is_valid():
            new_password = form.cleaned_data['new_password']
            request.user.set_password(new_password)
            request.user.save()
            messages.success(request, 'Password changed successfully.')
            return redirect('password_change')
        else:
            context = self.get_context_data(**kwargs)
            context['form'] = form
            return self.render_to_response(context)

class GenerateReportView(LoginRequiredMixin, RoleContextMixin, TemplateView):
    template_name = 'core/reports/report_generator.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = ReportFilterForm(user=self.request.user)
        return context
    
    def post(self, request, *args, **kwargs):
        if 'download' in request.POST:
            return self.generate_excel_report(request)
        return self.get(request, *args, **kwargs)
    
    def generate_excel_report(self, request):
        form = ReportFilterForm(request.POST, user=request.user)
        
        # Get filtered data
        students = Student.objects.all()
        
        # Apply user-specific filters
        user = request.user
        if user.groups.filter(name='Principal').exists():
            principal_profile = PrincipalProfile.objects.get(user=user)
            students = students.filter(school=principal_profile.school)
        elif user.groups.filter(name='BEO').exists():
            beo_profile = BEOProfile.objects.get(user=user)
            students = students.filter(school__taluka=beo_profile.taluka)
        elif user.groups.filter(name='DDPI').exists():
            ddpi_profile = DDPIProfile.objects.get(user=user)
            students = students.filter(school__taluka__district=ddpi_profile.district)
        
        # Apply form filters
        if form.is_valid():
            if form.cleaned_data['standard']:
                students = students.filter(standard=form.cleaned_data['standard'])
            if form.cleaned_data['school']:
                students = students.filter(school=form.cleaned_data['school'])
            if form.cleaned_data['taluka']:
                students = students.filter(school__taluka=form.cleaned_data['taluka'])
        
        # Get assignment data
        assignment = None
        if form.is_valid() and form.cleaned_data['assignment']:
            assignment = form.cleaned_data['assignment']
            students = students.filter(standard=assignment.standard)
        
        # Create Excel workbook
        wb = Workbook()
        
        if assignment:
            # Single assignment selected - create one worksheet
            ws = wb.active
            ws.title = f"{assignment.title[:20]}..."[:31] if len(assignment.title) > 20 else assignment.title
            self._create_assignment_worksheet(ws, students.filter(standard=assignment.standard), assignment)
        else:
            # No assignment selected - create worksheets for all assignments
            wb.remove(wb.active)  # Remove default sheet
            
            # Get all assignments filtered by user permissions and form filters
            assignments = Assignment.objects.all()
            
            # Apply user-specific filters to assignments (filter by standards available in user's scope)
            available_standards = set()
            if user.groups.filter(name='Principal').exists():
                principal_profile = PrincipalProfile.objects.get(user=user)
                school_students = Student.objects.filter(school=principal_profile.school)
                available_standards = set(school_students.values_list('standard', flat=True).distinct())
            elif user.groups.filter(name='BEO').exists():
                beo_profile = BEOProfile.objects.get(user=user)
                taluka_students = Student.objects.filter(school__taluka=beo_profile.taluka)
                available_standards = set(taluka_students.values_list('standard', flat=True).distinct())
            elif user.groups.filter(name='DDPI').exists():
                ddpi_profile = DDPIProfile.objects.get(user=user)
                district_students = Student.objects.filter(school__taluka__district=ddpi_profile.district)
                available_standards = set(district_students.values_list('standard', flat=True).distinct())
            
            if available_standards:
                assignments = assignments.filter(standard__in=available_standards)
            
            # Apply additional form filters to assignments
            if form.is_valid():
                if form.cleaned_data['standard']:
                    assignments = assignments.filter(standard=form.cleaned_data['standard'])
                if form.cleaned_data['subject']:
                    assignments = assignments.filter(subject=form.cleaned_data['subject'])
                if form.cleaned_data['start_date']:
                    assignments = assignments.filter(start_date__gte=form.cleaned_data['start_date'])
                if form.cleaned_data['end_date']:
                    assignments = assignments.filter(end_date__lte=form.cleaned_data['end_date'])
            
            assignments = assignments.distinct().order_by('standard', 'subject__name', 'title')
            
            if not assignments.exists():
                # Create a default worksheet if no assignments found
                ws = wb.create_sheet("No Assignments Found")
                ws.cell(row=1, column=1, value="No assignments found matching the selected criteria.")
            else:
                for assignment in assignments:
                    # Create worksheet name (max 31 chars for Excel)
                    sheet_name = f"Class{assignment.standard}-{assignment.subject.name[:10]}-{assignment.title[:10]}"
                    if len(sheet_name) > 31:
                        sheet_name = sheet_name[:31]
                    
                    ws = wb.create_sheet(sheet_name)
                    assignment_students = students.filter(standard=assignment.standard)
                    self._create_assignment_worksheet(ws, assignment_students, assignment)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"student_report_{timestamp}.xlsx"
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    def _create_assignment_worksheet(self, ws, students, assignment):
        """Helper method to create a worksheet for a specific assignment"""
        # Headers
        headers = ['Student Name', 'STS Number', 'Gender', 'Class', 'School', 'Taluka']
        
        for i, task in enumerate(assignment.tasks):
            headers.append(f'Task {i+1}: {task[:30]}...' if len(task) > 30 else f'Task {i+1}: {task}')
        
        # Apply header styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data rows
        for row, student in enumerate(students, 2):
            ws.cell(row=row, column=1, value=student.name)
            ws.cell(row=row, column=2, value=student.sts_number)
            ws.cell(row=row, column=3, value=student.get_gender_display())
            ws.cell(row=row, column=4, value=f'Class {student.standard}')
            ws.cell(row=row, column=5, value=student.school.name)
            ws.cell(row=row, column=6, value=student.school.taluka.name)
            
            for i, task in enumerate(assignment.tasks):
                try:
                    evaluation = TaskEvaluation.objects.get(
                        student=student,
                        assignment=assignment,
                        task_index=i
                    )
                    status = evaluation.status.upper()
                except TaskEvaluation.DoesNotExist:
                    status = 'UNSOLVED'
                
                ws.cell(row=row, column=7+i, value=status)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width